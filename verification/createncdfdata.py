#!/usr/bin/env python
# -*- coding: utf-8 -*-
from __future__ import print_function
from openpyxl import workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import numpy as np
from datetime import datetime, timedelta
import pytz
import  os, sys, argparse
from toolbox import MyTimeVars

import memstations

parser = argparse.ArgumentParser('createncdfdata.py')
parser.add_argument('input_file',help='path to .xlsx file')
parser.add_argument('out_file', help='Path to the netcdf timeseries file')

args  = parser.parse_args()
datafile=load_workbook(args.input_file)

list_data=[]
times=[]
stlist=[]
count=0
for row in datafile.active.iter_rows(min_row=1):
    count += 1
    dc=[]
    if count==1:
      for cell in row[1:]:
          try:
            stlist.append(memstations.name2code[cell.value])
          except KeyError:
            stlist.append(None)
            if (cell.value[0:3] in memstations.ignore):
                print("Ignoring", cell.value)
                pass
            else:
                print (cell.value[0:3])
                print("Not found", cell.value)
      continue
        

    for cell in row:
        if (cell.value==None):
            dc.append(np.nan)
        else:
            dc.append(cell.value)
    time = memstations.localtz.localize(datetime.strptime(dc[0], '%d/%m/%Y %H:%M'))
    timeUTC = time.astimezone(pytz.utc).replace(tzinfo=None)
    times.append(timeUTC)
    list_data.append(dc[1:])
#    print (time, timeUTC, dc[1:])

print("Got %d records from %s"%(count,args.input_file))
    
#    if count > 20:
#        break



tsfile = "MEMcoordinates.txt"
stationdic = memstations.getStations(tsfile)

stations_out=[]
stidxin={}

for i,st in enumerate(sorted(stlist)):
    if st is None:
        continue
    stidxin[st]=i
    stations_out.append(stationdic[st])


nst = len(stations_out)

ntimes = len(times)





times = np.array(times)
list_data = np.array(list_data, dtype=np.float32)



valmatr= np.float32(np.nan) * np.empty((ntimes,nst),dtype=np.float32)
for i,st in enumerate(stations_out):
    ist = stidxin[st.code]
    valmatr[:,i] = list_data[:,ist]

tsmatr=MyTimeVars.TsMatrix(times,stations_out,valmatr, "mg/m3")

tsmatr.to_nc(args.out_file)
print (args.out_file, "done!")







